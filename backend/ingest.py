import os, sys, hashlib, argparse, re
from pathlib import Path
from typing import List, Dict, Optional, Any, Tuple

import fitz
import docx
import openpyxl

from sentence_transformers import SentenceTransformer
import chromadb
from chromadb.config import Settings

BASE_DATA_DIR = Path(os.getenv("DATA_DIR", "/data/docs"))
BASE_CHROMA_DIR = os.getenv("CHROMA_DIR", "/data/chroma")
BASE_COLLECTION_NAME = os.getenv("CHROMA_COLLECTION_NAME", "docs")

EMBED_MODEL_NAME = os.getenv("EMBEDDING_MODEL_NAME", "BAAI/bge-m3")
EMBEDDING_DEVICE = os.getenv("EMBEDDING_DEVICE", "cpu")
HF_HOME = os.getenv("HF_HOME", "/workspace/hf")


def file_sha1(p: Path) -> str:
    h = hashlib.sha1()
    with open(p, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def chunk_text(text: str, max_chars: int = 1200, overlap: int = 150) -> List[str]:
    text = text.strip().replace("\r", "")
    chunks = []
    start = 0
    n = len(text)
    while start < n:
        end = min(start + max_chars, n)
        cut = text.rfind("\n", start, end)
        if cut == -1 or cut <= start + 200:
            cut = end
        seg = text[start:cut].strip()
        if seg:
            chunks.append(seg)
        if cut >= n:
            break
        start = max(cut - overlap, 0)
        if start == cut:
            start += 1
    return chunks


def extract_pdf(p: Path) -> List[Dict]:
    out = []
    with fitz.open(p) as doc:
        for i, page in enumerate(doc):
            txt = page.get_text("text")
            for j, ch in enumerate(chunk_text(txt)):
                out.append(
                    {
                        "text": ch,
                        "metadata": {
                            "path": str(p),
                            "type": "pdf",
                            "page": i + 1,
                            "chunk": j,
                        },
                    }
                )
    return out


def extract_docx(p: Path) -> List[Dict]:
    d = docx.Document(str(p))
    txt = "\n".join([para.text for para in d.paragraphs])
    return [
        {"text": ch, "metadata": {"path": str(p), "type": "docx"}}
        for ch in chunk_text(txt)
    ]


# =========================
# Excel ingest “tabular”
# =========================

PIVOT_BAD_HEADERS = {"etiquetas de fila", "etiquetas de columna", "total general", "suma de"}

KEY_PATTERNS = {
    "contenedor": re.compile(r"\bcontenedor\b|\bcontainer\b|\bcntr\b", re.I),
    "factura": re.compile(r"\bfactura\b|\binvoice\b", re.I),
    "pi": re.compile(r"\bpi\b|\bp\.?i\.?\b", re.I),
    "remision": re.compile(r"\bremisi[oó]n\b|\bremision\b|\brm\b", re.I),
    "modelo": re.compile(r"\bmodelo\b|\bmodel\b", re.I),
    "piezas": re.compile(r"\bpiezas\b|\bqty\b|\bpcs\b|\bpieces\b", re.I),
    "estatus": re.compile(r"\bestatus\b|\bstatus\b|\bestado\b", re.I),
    "anio": re.compile(r"\ba[nñ]o\b|\byear\b", re.I),
    "mes": re.compile(r"\bmes\b|\bmonth\b", re.I),
    "semana": re.compile(r"\bsemana\b|\bweek\b", re.I),
    "transporte": re.compile(r"\btransporte\b|\bcarrier\b|\btransport\b", re.I),
    "modalidad": re.compile(r"\bmodalidad\b|\bmode\b", re.I),
    "retailer": re.compile(r"\bretailer\b|\bcliente\b|\btienda\b", re.I),
}

IMPORTANT_ORDER = [
    "contenedor", "piezas", "factura", "pi", "remision",
    "modelo", "estatus", "anio", "mes", "semana",
    "transporte", "modalidad", "retailer",
]


def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    return str(h).strip()


def _norm_id(v: Any) -> str:
    s = str(v).strip()
    return s.upper()


def _is_pivot_like(headers: List[str]) -> bool:
    low = " ".join([h.lower() for h in headers if h])
    return any(bad in low for bad in PIVOT_BAD_HEADERS)


def _find_header_row(ws, max_scan: int = 40, min_nonempty: int = 5) -> Optional[int]:
    best_row = None
    best_score = 0.0

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        headers = [_norm_header(c) for c in row]
        nonempty = [h for h in headers if h]
        if len(nonempty) < min_nonempty:
            continue

        strings = sum(1 for c in row if isinstance(c, str) and str(c).strip())
        score = len(nonempty) + 0.2 * strings

        if _is_pivot_like(nonempty):
            score *= 0.2

        if score > best_score:
            best_score = score
            best_row = r_idx

    return best_row


def _pick_key_meta(headers: List[str], row_values: Tuple[Any, ...]) -> Dict[str, Any]:
    found: Dict[str, Any] = {}
    for h, v in zip(headers, row_values):
        if not h or v is None:
            continue
        hs = str(h).strip()
        if not hs:
            continue
        vs = str(v).strip()
        if not vs:
            continue

        for key, pat in KEY_PATTERNS.items():
            if pat.search(hs):
                if key in ("contenedor", "factura", "pi", "remision", "modelo"):
                    found[key] = _norm_id(vs)
                else:
                    found[key] = vs
    return found


def _build_row_text(
    ws_title: str,
    row_idx: int,
    headers: List[str],
    row_values: Tuple[Any, ...],
    key_meta: Dict[str, Any],
    max_extra: int = 12,
) -> str:
    parts: List[str] = [f"Hoja: {ws_title}", f"Fila: {row_idx}"]

    for k in IMPORTANT_ORDER:
        if k in key_meta:
            parts.append(f"{k}: {key_meta[k]}")

    extra = 0
    for h, v in zip(headers, row_values):
        if extra >= max_extra:
            break
        if not h or v is None:
            continue
        hs = str(h).strip()
        vs = str(v).strip()
        if not hs or not vs:
            continue

        if any(KEY_PATTERNS[key].search(hs) for key in KEY_PATTERNS):
            continue

        parts.append(f"{hs}: {vs}")
        extra += 1

    return " | ".join(parts)


def extract_xlsx(p: Path) -> List[Dict]:
    wb = openpyxl.load_workbook(str(p), data_only=True)
    out: List[Dict] = []

    for ws in wb.worksheets:
        header_row = _find_header_row(ws)
        if header_row is None:
            continue

        header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), None)
        if not header_values:
            continue

        headers = [_norm_header(c) for c in header_values]
        nonempty_headers = [h for h in headers if h]

        if len(nonempty_headers) < 5:
            continue
        if _is_pivot_like(nonempty_headers):
            continue

        data_start = header_row + 1

        for row_idx, row_values in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
            if row_values is None:
                continue

            has_any = any(v is not None and str(v).strip() != "" for v in row_values)
            if not has_any:
                continue

            key_meta = _pick_key_meta(headers, row_values)
            row_text = _build_row_text(ws.title, row_idx, headers, row_values, key_meta)

            meta = {
                "path": str(p),
                "type": "xlsx",
                "sheet": ws.title,
                "row": row_idx,
                "header_row": header_row,
                **key_meta,
            }

            out.append({"text": row_text, "metadata": meta})

    return out


def extract_txt(p: Path) -> List[Dict]:
    txt = Path(p).read_text(encoding="utf-8", errors="ignore")
    return [
        {"text": ch, "metadata": {"path": str(p), "type": "txt"}}
        for ch in chunk_text(txt)
    ]


def extract_any(p: Path) -> List[Dict]:
    ext = p.suffix.lower()
    if ext == ".pdf":
        return extract_pdf(p)
    if ext == ".docx":
        return extract_docx(p)
    if ext in (".xlsx", ".xlsm"):
        return extract_xlsx(p)
    if ext == ".txt":
        return extract_txt(p)
    return []


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ingesta de documentos a Chroma con soporte por área.")
    parser.add_argument(
        "--area",
        type=str,
        default=os.getenv("AREA"),
        help="Área a indexar (logistica, general, sistemas, etc.). Si se omite, modo global.",
    )
    return parser.parse_args()


_embed_model: Optional[SentenceTransformer] = None


def get_embed_model() -> SentenceTransformer:
    global _embed_model
    if _embed_model is None:
        print(f"[INGEST] Cargando modelo de embeddings: {EMBED_MODEL_NAME}")
        print(f"[INGEST] Dispositivo embeddings: {EMBEDDING_DEVICE}")
        print(f"[INGEST] HF_HOME (cache): {HF_HOME}")
        _embed_model = SentenceTransformer(
            EMBED_MODEL_NAME,
            device=EMBEDDING_DEVICE,
            cache_folder=HF_HOME,
        )
    return _embed_model


def get_collection(area: Optional[str]):
    if area:
        chroma_dir = os.path.join(BASE_CHROMA_DIR, area)
        collection_name = f"{BASE_COLLECTION_NAME}_{area}"
    else:
        chroma_dir = BASE_CHROMA_DIR
        collection_name = BASE_COLLECTION_NAME

    print(f"[INGEST] Inicializando Chroma en {chroma_dir}")
    client = chromadb.PersistentClient(
        path=chroma_dir,
        settings=Settings(allow_reset=False),
    )
    print(f"[INGEST] Colección: {collection_name}")
    return client.get_or_create_collection(collection_name)


def ingest_single_file(p: Path, area: Optional[str], model: SentenceTransformer, coll) -> int:
    print(f"[INGEST] {p}")
    docs = extract_any(p)
    if not docs:
        print(f"[SKIP] {p} (no se pudo extraer texto)")
        return 0

    # Re-index limpio por archivo (evita basura vieja)
    try:
        coll.delete(where={"path": str(p)})
    except Exception:
        pass

    base_id = file_sha1(p)[:12]
    ids, metadatas, texts = [], [], []

    for k, d in enumerate(docs):
        ids.append(f"{base_id}-{k}")
        meta = dict(d["metadata"]) if isinstance(d["metadata"], dict) else {}
        if area:
            meta["area"] = area
        metadatas.append(meta)
        texts.append(d["text"])

    embs = model.encode(texts, normalize_embeddings=True).tolist()
    coll.add(ids=ids, documents=texts, embeddings=embs, metadatas=metadatas)

    print(f"[OK] {p} → {len(ids)} chunks indexados")
    return len(ids)


def ingest_file_for_area(path: Path, area: Optional[str] = None) -> int:
    model = get_embed_model()
    coll = get_collection(area)
    return ingest_single_file(path, area, model, coll)


def main():
    args = parse_args()
    area = args.area

    data_dir = (BASE_DATA_DIR / area) if area else BASE_DATA_DIR

    if not data_dir.exists():
        print(f"[ERROR] No existe {data_dir}. Crea la carpeta y coloca documentos.", file=sys.stderr)
        sys.exit(1)

    print(f"[INFO] Área: {area if area else '(global sin área)'}")

    model = get_embed_model()
    coll = get_collection(area)

    files: List[Path] = []
    for root, _, fs in os.walk(data_dir):
        for name in fs:
            p = Path(root) / name
            if p.suffix.lower() in (".pdf", ".docx", ".xlsx", ".xlsm", ".txt"):
                files.append(p)

    if not files:
        print(f"[WARN] No se encontraron archivos soportados en {data_dir}.", file=sys.stderr)
        sys.exit(0)

    added = 0
    for p in sorted(files):
        added += ingest_single_file(p, area, model, coll)

    print(f"[DONE] Total chunks indexados: {added}")


if __name__ == "__main__":
    main()
