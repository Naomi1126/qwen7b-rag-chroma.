import os, sys, hashlib, argparse
from pathlib import Path
from typing import List, Dict, Optional

import fitz  # PyMuPDF
import docx
import openpyxl

from sentence_transformers import SentenceTransformer
import chromadb
from chromadb.config import Settings

BASE_DATA_DIR = Path(os.getenv("DATA_DIR", "/data/docs"))
BASE_CHROMA_DIR = os.getenv("CHROMA_DIR", "/data/chroma")
BASE_COLLECTION_NAME = os.getenv("CHROMA_COLLECTION_NAME", "docs")

EMBED_MODEL_NAME = os.getenv("EMBEDDING_MODEL_NAME", "BAAI/bge-m3")
EMBEDDING_DEVICE = os.getenv("EMBEDDING_DEVICE", "cpu")
HF_HOME = os.getenv("HF_HOME", "/workspace/hf")


def file_sha1(p: Path) -> str:
    h = hashlib.sha1()
    with open(p, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def chunk_text(text: str, max_chars: int = 1200, overlap: int = 150) -> List[str]:
    text = text.strip().replace("\r", "")
    chunks = []
    start = 0
    n = len(text)
    while start < n:
        end = min(start + max_chars, n)
        cut = text.rfind("\n", start, end)
        if cut == -1 or cut <= start + 200:
            cut = end
        seg = text[start:cut].strip()
        if seg:
            chunks.append(seg)
        if cut >= n:
            break
        start = max(cut - overlap, 0)
        if start == cut:
            start += 1
    return chunks


def extract_pdf(p: Path) -> List[Dict]:
    out = []
    with fitz.open(p) as doc:
        for i, page in enumerate(doc):
            txt = page.get_text("text")
            for j, ch in enumerate(chunk_text(txt)):
                out.append(
                    {
                        "text": ch,
                        "metadata": {
                            "path": str(p),
                            "type": "pdf",
                            "page": i + 1,
                            "chunk": j,
                        },
                    }
                )
    return out


def extract_docx(p: Path) -> List[Dict]:
    d = docx.Document(str(p))
    txt = "\n".join([para.text for para in d.paragraphs])
    return [
        {"text": ch, "metadata": {"path": str(p), "type": "docx"}}
        for ch in chunk_text(txt)
    ]


def extract_xlsx(p: Path) -> List[Dict]:
    wb = openpyxl.load_workbook(str(p), data_only=True)
    out: List[Dict] = []

    for ws in wb.worksheets:
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row is None:
            continue

        headers: List[str] = []
        for cell in first_row:
            headers.append("" if cell is None else str(cell).strip())

        for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            parts: List[str] = []
            for col_name, value in zip(headers, row_values):
                if not col_name or not col_name.strip():
                    continue
                if value is None:
                    continue
                val_str = str(value).strip()
                if val_str == "":
                    continue
                parts.append(f"{col_name}: {val_str}")

            if not parts:
                continue

            row_text = f"Hoja: {ws.title} | Fila: {row_idx} | " + " | ".join(parts)

            out.append(
                {
                    "text": row_text,
                    "metadata": {
                        "path": str(p),
                        "type": "xlsx",
                        "sheet": ws.title,
                        "row": row_idx,
                    },
                }
            )
    return out


def extract_txt(p: Path) -> List[Dict]:
    txt = Path(p).read_text(encoding="utf-8", errors="ignore")
    return [
        {"text": ch, "metadata": {"path": str(p), "type": "txt"}}
        for ch in chunk_text(txt)
    ]


def extract_any(p: Path) -> List[Dict]:
    ext = p.suffix.lower()
    if ext == ".pdf":
        return extract_pdf(p)
    if ext == ".docx":
        return extract_docx(p)
    if ext in (".xlsx", ".xlsm"):
        return extract_xlsx(p)
    if ext == ".txt":
        return extract_txt(p)
    return []


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ingesta de documentos a Chroma con soporte por área.")
    parser.add_argument(
        "--area",
        type=str,
        default=os.getenv("AREA"),
        help="Área a indexar (logistica, general, sistemas, etc.). Si se omite, modo global.",
    )
    return parser.parse_args()


_embed_model: Optional[SentenceTransformer] = None


def get_embed_model() -> SentenceTransformer:
    global _embed_model
    if _embed_model is None:
        print(f"[INGEST] Cargando modelo de embeddings: {EMBED_MODEL_NAME}")
        print(f"[INGEST] Dispositivo embeddings: {EMBEDDING_DEVICE}")
        print(f"[INGEST] HF_HOME (cache): {HF_HOME}")
        _embed_model = SentenceTransformer(
            EMBED_MODEL_NAME,
            device=EMBEDDING_DEVICE,
            cache_folder=HF_HOME,
        )
    return _embed_model


def get_collection(area: Optional[str]):
    if area:
        chroma_dir = os.path.join(BASE_CHROMA_DIR, area)
        collection_name = f"{BASE_COLLECTION_NAME}_{area}"
    else:
        chroma_dir = BASE_CHROMA_DIR
        collection_name = BASE_COLLECTION_NAME

    print(f"[INGEST] Inicializando Chroma en {chroma_dir}")
    client = chromadb.PersistentClient(
        path=chroma_dir,
        settings=Settings(allow_reset=False),
    )
    print(f"[INGEST] Colección: {collection_name}")
    return client.get_or_create_collection(collection_name)


def ingest_single_file(p: Path, area: Optional[str], model: SentenceTransformer, coll) -> int:
    print(f"[INGEST] {p}")
    docs = extract_any(p)
    if not docs:
        print(f"[SKIP] {p} (no se pudo extraer texto)")
        return 0

    base_id = file_sha1(p)[:12]
    ids, metadatas, texts = [], [], []

    for k, d in enumerate(docs):
        ids.append(f"{base_id}-{k}")
        meta = dict(d["metadata"]) if isinstance(d["metadata"], dict) else {}
        if area:
            meta["area"] = area
        metadatas.append(meta)
        texts.append(d["text"])

    # Re-index idempotente
    if ids:
        try:
            coll.delete(ids=ids)
        except Exception:
            pass

    embs = model.encode(texts, normalize_embeddings=True).tolist()
    coll.add(ids=ids, documents=texts, embeddings=embs, metadatas=metadatas)
    print(f"[OK] {p} → {len(ids)} chunks indexados")
    return len(ids)


def ingest_file_for_area(path: Path, area: Optional[str] = None) -> int:
    model = get_embed_model()
    coll = get_collection(area)
    return ingest_single_file(path, area, model, coll)


def main():
    args = parse_args()
    area = args.area

    data_dir = (BASE_DATA_DIR / area) if area else BASE_DATA_DIR

    if not data_dir.exists():
        print(f"[ERROR] No existe {data_dir}. Crea la carpeta y coloca documentos.", file=sys.stderr)
        sys.exit(1)

    print(f"[INFO] Área: {area if area else '(global sin área)'}")

    model = get_embed_model()
    coll = get_collection(area)

    files: List[Path] = []
    for root, _, fs in os.walk(data_dir):
        for name in fs:
            p = Path(root) / name
            if p.suffix.lower() in (".pdf", ".docx", ".xlsx", ".xlsm", ".txt"):
                files.append(p)

    if not files:
        print(f"[WARN] No se encontraron archivos soportados en {data_dir}.", file=sys.stderr)
        sys.exit(0)

    added = 0
    for p in sorted(files):
        added += ingest_single_file(p, area, model, coll)

    print(f"[DONE] Total chunks indexados: {added}")


if __name__ == "__main__":
    main()
